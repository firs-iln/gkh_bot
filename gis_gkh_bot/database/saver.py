from dataclasses import asdict
from pathlib import Path
from typing import NoReturn

import ujson
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.worksheet.worksheet import Worksheet

from parser.mkd import MKD
from parser.organization import Organization
from parser.room import Room
from parser.utils import BASE_DIR


class MKDDataSaver:
    def __init__(self, mkd: MKD, orgs: list[Organization], rooms: list[Room] = None):
        self.__wb = Workbook()
        self.__mkd_xl_sheet = self.__wb.create_sheet('МКД')
        self.__orgs_xl_sheet = self.__wb.create_sheet('Организации')
        self.__rooms_xl_sheet = self.__wb.create_sheet('Помещения')
        del self.__wb['Sheet']
        self.__mkd = mkd
        self.__orgs = orgs
        self.__rooms = rooms

    async def save_mkd_to_excel(self):
        logger.info('Saving MKD to Excel')
        await self.__write_sheets_headers()
        await self.__write_mkd_data()
        await self.__write_orgs_data()
        if self.__rooms:
            await self.__write_rooms_data()
        self.__apply_styles()
        filepath = Path(BASE_DIR / 'reports' / f'МКД_{self.__mkd.cad_id}_{self.__mkd.address.replace("/", "_")}.xlsx').resolve()
        # filepath.touch(0o777)
        # with open(filepath, 'w+'):
        #     pass
        self.__wb.save(filepath)

    async def __write_sheets_headers(self) -> NoReturn:
        self.__mkd_xl_sheet.append(
            ['ID МКД', 'Адрес МКД:', 'Кадастровый номер:', 'Идентификационный код адреса:', 'Общ.площадь МКД',
             'Общ.площадь КВ', 'Год постройки', 'Способ управления:', 'ИНН УО', 'ИНН РСО', 'Ссылка 1', 'Ссылка 2',
             'Ссылка 3', 'Код субьекта', 'Индекс', 'Город / н.п.', 'Улица', 'Дом', 'КадНомМКД', 'Геокоординаты']
        )
        self.__orgs_xl_sheet.append(
            ['ИНН организации', 'Статус исп-ля', 'Наименование организации', 'Субьект РФ', 'ОГРН',
             'Дата гос. регистрации', 'КПП', 'E-mail', 'Контактный телефон', 'Тел. диспетчерской', 'ФИО руководителя',
             'Должность руководителя', 'Все Функции', 'Ссылка', 'Статус', 'Наименование краткое', 'ОГРН', 'ФИО ЕИО',
             'Должность ЕИО', 'Телефон', 'Email', 'Ссылка 5']
        )
        if self.__rooms:
            self.__rooms_xl_sheet.append(
                ['ID помещ', 'ID МКД', 'Номер', 'КадНом', 'УстНом', 'Площадь, м2', 'Статус', 'Жилая площь', 'Комнат',
                 'Подъезд', 'Аварийное', 'Росреестр', 'Адрес', 'Квартира', 'Кадастровый Номер', 'Код ФИАС ГАР',
                 'Площадь', 'Уровень', 'ID МКД']
            )
        else:
            self.__rooms_xl_sheet.append(
                ['Данные о помещениях будут доступны через некоторое время, вы будете уведомлены об этом'])

    async def __write_mkd_data(self) -> NoReturn:
        data = asdict(self.__mkd)
        try:
            data['total_area'] = float(str(data['total_area']).replace(',', '.'))
            data['residential_square'] = float(str(data['residential_square']).replace(',', '.'))
        except ValueError:
            pass
        logger.debug(data)
        self.__mkd_xl_sheet.append(tuple(data.values()))

    async def __write_orgs_data(self) -> NoReturn:
        for org in self.__orgs:
            data = asdict(org)
            self.__orgs_xl_sheet.append(tuple(data.values()))

    async def __write_rooms_data(self) -> NoReturn:
        rooms = self.__rooms
        for room in rooms:
            data = asdict(room)
            try:
                data['total_area'] = float(str(data['total_area']).replace(',', '.'))
                data['residential_square'] = float(str(data['residential_square']).replace(',', '.'))
            except ValueError:
                pass
            self.__rooms_xl_sheet.append(tuple(data.values()))

    def __apply_styles(self) -> NoReturn:
        self.__set_widths()
        for sheet in (self.__mkd_xl_sheet, self.__orgs_xl_sheet, self.__rooms_xl_sheet):
            self.__set_sheet_col_alignment_and_borders(sheet)
            self.__set_sheet_header_fill_and_alignment(sheet)
        self.__set_orgs_funcs_col_style()
        self.__set_font_sizes()

    def __set_widths(self) -> NoReturn:
        with open('widths.json', 'r') as f:
            widths = ujson.load(f)
        for sheet in (self.__mkd_xl_sheet, self.__orgs_xl_sheet, self.__rooms_xl_sheet):
            for dim, width in widths[sheet.title].items():
                sheet.column_dimensions[dim].width = float(width)

    @staticmethod
    def __set_sheet_col_alignment_and_borders(sheet: Worksheet) -> NoReturn:
        alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        thin_border = Border(
            left=Side(border_style=BORDER_THIN, color='00000000'),
            right=Side(border_style=BORDER_THIN, color='00000000'),
            top=Side(border_style=BORDER_THIN, color='00000000'),
            bottom=Side(border_style=BORDER_THIN, color='00000000')
        )
        for col in sheet.columns:
            for cell in col:
                if sheet.title == 'Помещения' and cell.column_letter in ('C', 'F', 'G', 'H', 'I', 'K', 'L'):
                    cell.alignment = Alignment(vertical='top', horizontal='center', wrap_text=True)
                else:
                    cell.alignment = alignment
                cell.border = thin_border

    @staticmethod
    def __set_sheet_header_fill_and_alignment(sheet: Worksheet) -> NoReturn:
        for row in sheet.iter_rows(1, 1):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="00CCFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for row in sheet.iter_rows(1, 1, min_col=14):
            for cell in row:
                cell.fill = PatternFill('solid', fgColor='00FFCC99')
        sheet.freeze_panes = 'A2'

    def __set_orgs_funcs_col_style(self) -> NoReturn:
        for col in self.__orgs_xl_sheet.iter_cols(min_col=13, max_col=13, min_row=2):
            for cell in col:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)
                cell.font = Font(sz=8)

    def __set_font_sizes(self) -> NoReturn:
        for col in self.__orgs_xl_sheet.iter_cols(min_col=13, max_col=14, min_row=2):
            for cell in col:
                cell.font = Font(sz=8)
                cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=False)

        for col in self.__orgs_xl_sheet.iter_cols(min_col=22, max_col=22, min_row=2):
            for cell in col:
                cell.font = Font(sz=8)
                cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=False)

        for col in self.__mkd_xl_sheet.iter_cols(min_col=11, max_col=13, min_row=2):
            for cell in col:
                cell.font = Font(sz=8)
                cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=False)

        for col in self.__mkd_xl_sheet.iter_cols(min_col=20, max_col=20, min_row=2):
            for cell in col:
                cell.font = Font(sz=8)
                cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=False)

        for col in self.__rooms_xl_sheet.iter_cols(min_col=13, max_col=13, min_row=2):
            for cell in col:
                cell.font = Font(sz=8)
                cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=False)

        for col in self.__rooms_xl_sheet.iter_cols(min_col=5, max_col=5, min_row=2):
            for cell in col:
                cell.font = Font(sz=8)
                cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=False)

        for col in self.__rooms_xl_sheet.iter_cols(min_col=16, max_col=16, min_row=2):
            for cell in col:
                cell.font = Font(sz=8)
                cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=False)
